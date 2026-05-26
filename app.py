import os
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, session, send_file
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import inspect, or_, text
from werkzeug.utils import secure_filename
from pathlib import Path
import csv
import io
import json
import socket
from datetime import datetime, time
import openpyxl
from markupsafe import Markup

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "change-this-secret-for-production")

BASE = Path(__file__).resolve().parent
UPLOADS_DIR = BASE / "static" / "uploads"
ALLOWED_IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".gif"}

DATABASE_URL = os.environ.get("DATABASE_URL")
if DATABASE_URL:
    if DATABASE_URL.startswith("postgres://"):
        DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)
else:
    DATABASE_URL = f"sqlite:///{BASE / 'data.db'}"

app.config["SQLALCHEMY_DATABASE_URI"] = DATABASE_URL
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

db = SQLAlchemy(app)

LANGUAGES = {
    "zh": "中文",
    "en": "English",
    "ja": "日本語",
}

TRANSLATIONS = {
    "zh": {
        "app_name": "库存管理与收银系统",
        "home": "返回首页",
        "inventory": "库存管理",
        "cashier": "收银结账",
        "members": "会员管理",
        "sales": "销售记录",
        "customer_display": "顾客显示屏",
        "scan": "扫码",
        "start_scan": "开始扫码",
        "stop_scan": "停止扫码",
        "add": "加入",
        "save": "保存",
        "delete": "删除",
        "remove": "移除",
        "checkout": "结账",
        "print_receipt": "打印小票",
        "cancel_cart": "取消购物车",
        "barcode": "条形码",
        "product": "商品",
        "product_name": "商品名称",
        "image": "图片",
        "price": "单价",
        "stock": "库存",
        "qty": "数量",
        "subtotal": "小计",
        "action": "操作",
        "original_total": "原总价",
        "payable": "折后价",
        "total": "总计",
        "paid_amount": "实收金额",
        "payment_method": "付款方式",
        "cash": "现金",
        "card": "刷卡",
        "mobile_pay": "移动支付",
        "other": "其他",
        "barcode_entry": "扫码 / 输入条码",
        "barcode_placeholder": "扫码或输入条形码",
        "product_search": "商品搜索",
        "product_search_placeholder": "搜索商品名称或条码",
        "search": "搜索",
        "tap_product_hint": "可直接点选商品加入购物车。",
        "member_discount": "会员折扣",
        "member_id": "会员卡号",
        "verify": "验证",
        "clear": "清除",
        "no_member_discount": "当前未应用会员折扣。",
        "cart": "购物车",
        "empty_cart": "购物车为空",
        "waiting_for_cart": "等待收银录入商品",
        "last_updated": "最后更新",
        "paid": "已结账",
        "checkout_complete": "结账完成",
        "thank_you": "感谢购买",
        "stock_after_checkout": "结账后库存将自动减少。",
        "sales_order": "销售单",
        "time": "时间",
        "member": "会员",
        "orders": "订单数",
        "sold_items": "销售件数",
        "details": "明细",
        "view_date": "查看日期",
        "download_excel": "下载 Excel",
        "download_csv": "下载 CSV",
        "no_sales": "当天暂无销售记录",
        "new_member": "新增会员",
        "member_list": "会员列表",
        "member_name": "会员姓名",
        "discount": "折扣(%)",
        "save_member": "保存会员",
        "no_members": "暂无会员",
        "edit_product": "录入/修改商品",
        "save_product": "保存商品",
        "import_export": "导入 / 导出",
        "import_export_hint": "上传 CSV 或 Excel 文件后，系统会将商品信息导入当前库存；导出将生成带商品信息的文件。",
        "choose_file": "选择文件",
        "import_inventory": "导入库存",
        "export_csv": "导出 CSV",
        "export_excel": "导出 Excel",
        "current_inventory": "当前库存",
        "no_products": "暂无商品，请先录入",
        "scanner": "扫码器",
        "home_intro": "这是一个简单的库存管理与POS收银系统。你可以：",
        "home_feature_inventory": "在库存管理页面录入、修改或删除商品",
        "home_feature_cashier": "在收银页面扫码或输入条形码，实时显示商品信息、单价和总价",
        "home_feature_stock": "结账后自动更新库存",
        "lan_hint": "如果你希望用手机扫码，请在同一局域网下用手机浏览器访问：",
        "lan_help": "如果打不开，请确认电脑和手机在同一 Wi-Fi、端口未被防火墙阻断。",
        "none": "无",
        "no_image": "无图",
    },
    "en": {
        "app_name": "Inventory POS System",
        "home": "Home",
        "inventory": "Inventory",
        "cashier": "Cashier",
        "members": "Members",
        "sales": "Sales Records",
        "customer_display": "Customer Display",
        "scan": "Scan",
        "start_scan": "Start Scan",
        "stop_scan": "Stop Scan",
        "add": "Add",
        "save": "Save",
        "delete": "Delete",
        "remove": "Remove",
        "checkout": "Checkout",
        "print_receipt": "Print Receipt",
        "cancel_cart": "Cancel Cart",
        "barcode": "Barcode",
        "product": "Product",
        "product_name": "Product Name",
        "image": "Image",
        "price": "Price",
        "stock": "Stock",
        "qty": "Qty",
        "subtotal": "Subtotal",
        "action": "Action",
        "original_total": "Original Total",
        "payable": "Payable",
        "total": "Total",
        "paid_amount": "Paid Amount",
        "payment_method": "Payment Method",
        "cash": "Cash",
        "card": "Card",
        "mobile_pay": "Mobile Pay",
        "other": "Other",
        "barcode_entry": "Scan / Enter Barcode",
        "barcode_placeholder": "Scan or enter barcode",
        "product_search": "Product Search",
        "product_search_placeholder": "Search name or barcode",
        "search": "Search",
        "tap_product_hint": "Tap a product to add it to the cart.",
        "member_discount": "Member Discount",
        "member_id": "Member ID",
        "verify": "Verify",
        "clear": "Clear",
        "no_member_discount": "No member discount applied.",
        "cart": "Cart",
        "empty_cart": "Cart is empty",
        "waiting_for_cart": "Waiting for items",
        "last_updated": "Last Updated",
        "paid": "Paid",
        "checkout_complete": "Checkout complete",
        "thank_you": "Thank you",
        "stock_after_checkout": "Stock will be reduced after checkout.",
        "sales_order": "Sale",
        "time": "Time",
        "member": "Member",
        "orders": "Orders",
        "sold_items": "Items Sold",
        "details": "Details",
        "view_date": "View Date",
        "download_excel": "Download Excel",
        "download_csv": "Download CSV",
        "no_sales": "No sales records for this date",
        "new_member": "New Member",
        "member_list": "Member List",
        "member_name": "Member Name",
        "discount": "Discount (%)",
        "save_member": "Save Member",
        "no_members": "No members",
        "edit_product": "Add / Edit Product",
        "save_product": "Save Product",
        "import_export": "Import / Export",
        "import_export_hint": "Upload a CSV or Excel file to import products; export will download product data.",
        "choose_file": "Choose File",
        "import_inventory": "Import Inventory",
        "export_csv": "Export CSV",
        "export_excel": "Export Excel",
        "current_inventory": "Current Inventory",
        "no_products": "No products yet",
        "scanner": "Scanner",
        "home_intro": "A simple inventory and POS cashier system. You can:",
        "home_feature_inventory": "Add, edit, or delete products in inventory",
        "home_feature_cashier": "Scan or enter barcodes at checkout and see prices in real time",
        "home_feature_stock": "Automatically update stock after checkout",
        "lan_hint": "For mobile scanning on the same Wi-Fi, open:",
        "lan_help": "If it does not open, check Wi-Fi and firewall settings.",
        "none": "None",
        "no_image": "No Image",
    },
    "ja": {
        "app_name": "在庫管理・レジシステム",
        "home": "ホーム",
        "inventory": "在庫管理",
        "cashier": "レジ会計",
        "members": "会員管理",
        "sales": "売上記録",
        "customer_display": "お客様表示",
        "scan": "スキャン",
        "start_scan": "スキャン開始",
        "stop_scan": "スキャン停止",
        "add": "追加",
        "save": "保存",
        "delete": "削除",
        "remove": "削除",
        "checkout": "会計",
        "print_receipt": "レシート印刷",
        "cancel_cart": "カート取消",
        "barcode": "バーコード",
        "product": "商品",
        "product_name": "商品名",
        "image": "画像",
        "price": "単価",
        "stock": "在庫",
        "qty": "数量",
        "subtotal": "小計",
        "action": "操作",
        "original_total": "元合計",
        "payable": "割引後",
        "total": "合計",
        "paid_amount": "支払額",
        "payment_method": "支払方法",
        "cash": "現金",
        "card": "カード",
        "mobile_pay": "モバイル決済",
        "other": "その他",
        "barcode_entry": "スキャン / バーコード入力",
        "barcode_placeholder": "バーコードをスキャンまたは入力",
        "product_search": "商品検索",
        "product_search_placeholder": "商品名またはバーコードを検索",
        "search": "検索",
        "tap_product_hint": "商品をタップしてカートに追加できます。",
        "member_discount": "会員割引",
        "member_id": "会員番号",
        "verify": "確認",
        "clear": "クリア",
        "no_member_discount": "会員割引は適用されていません。",
        "cart": "カート",
        "empty_cart": "カートは空です",
        "waiting_for_cart": "商品入力待ち",
        "last_updated": "最終更新",
        "paid": "会計済み",
        "checkout_complete": "お会計が完了しました",
        "thank_you": "ありがとうございました",
        "stock_after_checkout": "会計後に在庫が自動で減ります。",
        "sales_order": "売上",
        "time": "時間",
        "member": "会員",
        "orders": "注文数",
        "sold_items": "販売数",
        "details": "明細",
        "view_date": "日付表示",
        "download_excel": "Excel ダウンロード",
        "download_csv": "CSV ダウンロード",
        "no_sales": "この日の売上記録はありません",
        "new_member": "新規会員",
        "member_list": "会員一覧",
        "member_name": "会員名",
        "discount": "割引(%)",
        "save_member": "会員保存",
        "no_members": "会員はありません",
        "edit_product": "商品登録 / 編集",
        "save_product": "商品保存",
        "import_export": "インポート / エクスポート",
        "import_export_hint": "CSV または Excel をアップロードして商品を取り込み、商品データを出力できます。",
        "choose_file": "ファイル選択",
        "import_inventory": "在庫インポート",
        "export_csv": "CSV 出力",
        "export_excel": "Excel 出力",
        "current_inventory": "現在の在庫",
        "no_products": "商品がありません",
        "scanner": "スキャナー",
        "home_intro": "シンプルな在庫管理・POS レジシステムです。できること：",
        "home_feature_inventory": "在庫管理で商品の登録・編集・削除",
        "home_feature_cashier": "レジでバーコードをスキャンまたは入力し、価格をリアルタイム表示",
        "home_feature_stock": "会計後に在庫を自動更新",
        "lan_hint": "同じ Wi-Fi のスマホでスキャンする場合はこちら：",
        "lan_help": "開けない場合は Wi-Fi とファイアウォールを確認してください。",
        "none": "なし",
        "no_image": "画像なし",
    },
}

PAYMENT_METHOD_KEYS = {
    "现金": "cash",
    "刷卡": "card",
    "移动支付": "mobile_pay",
    "其他": "other",
    "未记录": "none",
}

class Product(db.Model):
    __tablename__ = "products"
    barcode = db.Column(db.String(128), primary_key=True)
    name = db.Column(db.String(256), nullable=False)
    price = db.Column(db.Float, nullable=False, default=0.0)
    stock = db.Column(db.Integer, nullable=False, default=0)
    image = db.Column(db.String(512), nullable=True)

    def to_dict(self):
        return {
            "barcode": self.barcode,
            "name": self.name,
            "price": self.price,
            "stock": self.stock,
            "image": self.image,
        }

class Member(db.Model):
    __tablename__ = "members"
    member_id = db.Column(db.String(128), primary_key=True)
    name = db.Column(db.String(256), nullable=False)
    discount = db.Column(db.Integer, nullable=False, default=0)

    def to_dict(self):
        return {
            "member_id": self.member_id,
            "name": self.name,
            "discount": self.discount,
        }


class Sale(db.Model):
    __tablename__ = "sales"
    id = db.Column(db.Integer, primary_key=True)
    created_at = db.Column(db.DateTime, nullable=False, default=datetime.now)
    member_id = db.Column(db.String(128), nullable=True)
    member_name = db.Column(db.String(256), nullable=True)
    discount = db.Column(db.Integer, nullable=False, default=0)
    payment_method = db.Column(db.String(64), nullable=False, default="未记录")
    total = db.Column(db.Float, nullable=False, default=0.0)
    payable = db.Column(db.Float, nullable=False, default=0.0)
    items = db.relationship("SaleItem", backref="sale", cascade="all, delete-orphan", lazy=True)


class SaleItem(db.Model):
    __tablename__ = "sale_items"
    id = db.Column(db.Integer, primary_key=True)
    sale_id = db.Column(db.Integer, db.ForeignKey("sales.id"), nullable=False)
    barcode = db.Column(db.String(128), nullable=False)
    name = db.Column(db.String(256), nullable=False)
    price = db.Column(db.Float, nullable=False, default=0.0)
    qty = db.Column(db.Integer, nullable=False, default=0)
    subtotal = db.Column(db.Float, nullable=False, default=0.0)


class CustomerDisplayState(db.Model):
    __tablename__ = "customer_display_state"
    id = db.Column(db.Integer, primary_key=True)
    payload = db.Column(db.Text, nullable=False, default="{}")
    updated_at = db.Column(db.DateTime, nullable=False, default=datetime.now)


def ensure_directories():
    UPLOADS_DIR.mkdir(parents=True, exist_ok=True)


def ensure_schema():
    inspector = inspect(db.engine)
    if "sales" not in inspector.get_table_names():
        return
    sale_columns = {column["name"] for column in inspector.get_columns("sales")}
    if "payment_method" not in sale_columns:
        db.session.execute(text("ALTER TABLE sales ADD COLUMN payment_method VARCHAR(64) NOT NULL DEFAULT '未记录'"))
        db.session.commit()


with app.app_context():
    ensure_directories()
    db.create_all()
    ensure_schema()


def get_language():
    lang = session.get("lang", "zh")
    return lang if lang in LANGUAGES else "zh"


def translate(key):
    lang = get_language()
    return TRANSLATIONS.get(lang, TRANSLATIONS["zh"]).get(key, TRANSLATIONS["zh"].get(key, key))


def format_jpy(value):
    try:
        return str(int(round(float(value or 0))))
    except (TypeError, ValueError):
        return "0"


app.jinja_env.filters["jpy"] = format_jpy


def language_switch():
    current = get_language()
    back = request.full_path if request.query_string else request.path
    links = []
    for code, label in LANGUAGES.items():
        active = " active" if code == current else ""
        href = url_for("set_language", lang=code, next=back)
        links.append(f'<a class="lang-link{active}" href="{href}">{label}</a>')
    return Markup('<div class="language-switch">' + "".join(links) + "</div>")


@app.context_processor
def inject_language_tools():
    return {
        "t": translate,
        "current_lang": get_language,
        "language_switch": language_switch,
    }


@app.route("/language/<lang>")
def set_language(lang):
    if lang in LANGUAGES:
        session["lang"] = lang
    target = request.args.get("next") or url_for("index")
    if not target.startswith("/"):
        target = url_for("index")
    return redirect(target)


def load_inventory():
    return [product.to_dict() for product in Product.query.order_by(Product.barcode).all()]


def save_inventory(data):
    Product.query.delete()
    for item in data:
        try:
            product = Product(
                barcode=str(item.get("barcode", "") or "").strip(),
                name=str(item.get("name", "") or "").strip(),
                price=float(item.get("price", 0) or 0),
                stock=int(item.get("stock", 0) or 0),
                image=str(item.get("image", "") or None) if item.get("image") else None,
            )
        except (ValueError, TypeError):
            continue
        db.session.add(product)
    db.session.commit()


def load_members():
    return [member.to_dict() for member in Member.query.order_by(Member.member_id).all()]


def save_members(data):
    Member.query.delete()
    for item in data:
        try:
            member = Member(
                member_id=str(item.get("member_id", "") or "").strip(),
                name=str(item.get("name", "") or "").strip(),
                discount=int(item.get("discount", 0) or 0),
            )
        except (ValueError, TypeError):
            continue
        db.session.add(member)
    db.session.commit()


def find_product(barcode):
    return Product.query.get(barcode)


def find_member(member_id):
    return Member.query.get(member_id)


def allowed_image(filename):
    suffix = Path(filename).suffix.lower()
    return suffix in ALLOWED_IMAGE_EXTENSIONS


def save_image_file(file_storage, barcode):
    if not file_storage:
        return None
    filename = secure_filename(file_storage.filename)
    if not filename:
        return None
    suffix = Path(filename).suffix.lower()
    if suffix not in ALLOWED_IMAGE_EXTENSIONS:
        return None
    ensure_directories()
    target_name = f"{barcode}{suffix}"
    target_path = UPLOADS_DIR / target_name
    file_storage.save(target_path)
    return f"uploads/{target_name}"


def update_product(product):
    existing = Product.query.get(product["barcode"])
    if existing:
        existing.name = product["name"]
        existing.price = product["price"]
        existing.stock = product["stock"]
        existing.image = product.get("image")
    else:
        existing = Product(
            barcode=product["barcode"],
            name=product["name"],
            price=product["price"],
            stock=product["stock"],
            image=product.get("image"),
        )
        db.session.add(existing)
    db.session.commit()


def update_member(member):
    existing = Member.query.get(member["member_id"])
    if existing:
        existing.name = member["name"]
        existing.discount = member["discount"]
    else:
        existing = Member(
            member_id=member["member_id"],
            name=member["name"],
            discount=member["discount"],
        )
        db.session.add(existing)
    db.session.commit()


def get_local_ip():
    try:
        with socket.socket(socket.AF_INET, socket.SOCK_DGRAM) as s:
            s.connect(("8.8.8.8", 80))
            return s.getsockname()[0]
    except Exception:
        return "127.0.0.1"


def get_cart():
    return session.get("cart", {})


def save_cart(cart):
    session["cart"] = cart


def cart_items():
    cart = get_cart()
    items = []
    total = 0.0
    for barcode, qty in cart.items():
        product = find_product(barcode)
        if not product:
            continue
        subtotal = product.price * qty
        items.append({
            "barcode": barcode,
            "name": product.name,
            "price": product.price,
            "qty": qty,
            "subtotal": subtotal,
            "image": product.image,
        })
        total += subtotal
    return items, total


def cart_payload():
    items, total = cart_items()
    return {
        "items": items,
        "total": round(total),
    }


def display_payload(items=None, total=0, status="active", checkout_id=None):
    payload = {
        "items": items or [],
        "total": round(total or 0),
        "status": status,
        "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    if checkout_id is not None:
        payload["checkout_id"] = checkout_id
    return payload


def save_customer_display(payload):
    state = db.session.get(CustomerDisplayState, 1)
    if not state:
        state = CustomerDisplayState(id=1)
        db.session.add(state)
    state.payload = json.dumps(payload, ensure_ascii=False)
    state.updated_at = datetime.now()
    db.session.commit()
    return payload


def sync_customer_display_from_cart(status="active"):
    items, total = cart_items()
    return save_customer_display(display_payload(items, total, status))


def get_customer_display_payload():
    state = db.session.get(CustomerDisplayState, 1)
    if not state:
        return display_payload(status="idle")
    try:
        payload = json.loads(state.payload or "{}")
    except json.JSONDecodeError:
        return display_payload(status="idle")
    payload.setdefault("items", [])
    payload.setdefault("total", 0)
    payload.setdefault("status", "idle")
    payload.setdefault("checkout_id", None)
    payload.setdefault("updated_at", state.updated_at.strftime("%Y-%m-%d %H:%M:%S"))
    if payload.get("status") == "paid" and (datetime.now() - state.updated_at).total_seconds() > 7:
        return save_customer_display(display_payload(status="idle"))
    return payload


def calculate_payable(total, member):
    discount = member.discount if member else 0
    return round(total * (1 - discount / 100))


def normalize_payment_method(value):
    allowed_methods = {"现金", "刷卡", "移动支付", "其他"}
    method = str(value or "").strip()
    return method if method in allowed_methods else "其他"


def payment_method_label(value):
    key = PAYMENT_METHOD_KEYS.get(value)
    return translate(key) if key else (value or translate("none"))


def parse_report_date(value):
    if value:
        try:
            return datetime.strptime(value, "%Y-%m-%d").date()
        except ValueError:
            pass
    return datetime.now().date()


def sales_for_date(report_date):
    start = datetime.combine(report_date, time.min)
    end = datetime.combine(report_date, time.max)
    return (
        Sale.query
        .filter(Sale.created_at >= start, Sale.created_at <= end)
        .order_by(Sale.created_at.desc(), Sale.id.desc())
        .all()
    )


def sale_rows(sales):
    rows = []
    for sale in sales:
        for item in sale.items:
            rows.append({
                "sale_id": sale.id,
                "created_at": sale.created_at.strftime("%Y-%m-%d %H:%M:%S"),
                "member_id": sale.member_id or "",
                "member_name": sale.member_name or "",
                "discount": sale.discount,
                "payment_method": payment_method_label(sale.payment_method or "未记录"),
                "barcode": item.barcode,
                "name": item.name,
                "price": item.price,
                "qty": item.qty,
                "subtotal": item.subtotal,
                "order_total": sale.total,
                "order_payable": sale.payable,
            })
    return rows


def parse_inventory_file(file_storage):
    filename = file_storage.filename or ""
    suffix = Path(filename).suffix.lower()
    if suffix == ".csv":
        text = io.TextIOWrapper(file_storage.stream, encoding="utf-8")
        reader = csv.DictReader(text)
        products = []
        for row in reader:
            if not row.get("barcode"):
                continue
            try:
                products.append({
                    "barcode": str(row.get("barcode", "")).strip(),
                    "name": str(row.get("name", "")).strip(),
                    "price": float(row.get("price", 0)),
                    "stock": int(row.get("stock", 0)),
                    "image": str(row.get("image", "")).strip() or None,
                })
            except ValueError:
                raise ValueError("CSV 文件中的价格或库存格式错误")
        return products
    elif suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        workbook = openpyxl.load_workbook(file_storage, data_only=True)
        sheet = workbook.active
        headers = [str(cell.value).strip().lower() if cell.value else "" for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        products = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            row_data = dict(zip(headers, row))
            try:
                products.append({
                    "barcode": str(row_data.get("barcode", "")).strip(),
                    "name": str(row_data.get("name", "")).strip(),
                    "price": float(row_data.get("price", 0) or 0),
                    "stock": int(row_data.get("stock", 0) or 0),
                    "image": str(row_data.get("image", "")).strip() or None,
                })
            except ValueError:
                raise ValueError("Excel 文件中的价格或库存格式错误")
        return products
    else:
        raise ValueError("仅支持 CSV 或 Excel 文件")


@app.route("/")
def index():
    return render_template("index.html", host=request.host, local_ip=get_local_ip())


@app.route("/manage")
def manage():
    inventory = load_inventory()
    return render_template("manage.html", inventory=inventory)


@app.route("/manage/add", methods=["POST"])
def manage_add():
    barcode = request.form.get("barcode", "").strip()
    name = request.form.get("name", "").strip()
    price = request.form.get("price", "").strip()
    stock = request.form.get("stock", "").strip()
    image_file = request.files.get("image")

    if not barcode or not name or not price or not stock:
        flash("请填写完整商品信息", "error")
        return redirect(url_for("manage"))
    try:
        price = float(price)
        stock = int(stock)
    except ValueError:
        flash("价格必须是数字，库存必须是整数", "error")
        return redirect(url_for("manage"))
    if price < 0 or stock < 0:
        flash("价格和库存不能为负数", "error")
        return redirect(url_for("manage"))

    existing = find_product(barcode)
    image_path = existing.image if existing else None
    uploaded_image = save_image_file(image_file, barcode)
    if uploaded_image:
        image_path = uploaded_image
    product = {
        "barcode": barcode,
        "name": name,
        "price": price,
        "stock": stock,
        "image": image_path,
    }
    update_product(product)
    flash("商品已保存", "success")
    return redirect(url_for("manage"))


@app.route("/manage/delete/<barcode>", methods=["POST"])
def manage_delete(barcode):
    product = Product.query.get(barcode)
    if product and product.image:
        image_path = BASE / "static" / product.image
        if image_path.exists():
            try:
                image_path.unlink()
            except Exception:
                pass
    if product:
        db.session.delete(product)
        db.session.commit()
    flash("商品已删除", "success")
    return redirect(url_for("manage"))


@app.route("/manage/import", methods=["POST"])
def manage_import():
    inventory_file = request.files.get("inventory_file")
    if not inventory_file or not inventory_file.filename:
        flash("请上传 CSV 或 Excel 文件", "error")
        return redirect(url_for("manage"))
    try:
        products = parse_inventory_file(inventory_file)
        for product in products:
            update_product(product)
        flash(f"已导入 {len(products)} 条商品", "success")
    except ValueError as exc:
        flash(str(exc), "error")
    return redirect(url_for("manage"))


@app.route("/manage/export/<fmt>")
def manage_export(fmt):
    inventory = load_inventory()
    if fmt == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["barcode", "name", "price", "stock", "image"])
        for product in inventory:
            writer.writerow([
                product.get("barcode", ""),
                product.get("name", ""),
                product.get("price", ""),
                product.get("stock", ""),
                product.get("image", ""),
            ])
        data = output.getvalue().encode("utf-8-sig")
        return send_file(
            io.BytesIO(data),
            mimetype="text/csv",
            as_attachment=True,
            download_name="inventory.csv",
        )
    elif fmt in {"xlsx", "xlsm", "xltx", "xltm"}:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["barcode", "name", "price", "stock", "image"])
        for product in inventory:
            sheet.append([
                product.get("barcode", ""),
                product.get("name", ""),
                product.get("price", ""),
                product.get("stock", ""),
                product.get("image", ""),
            ])
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="inventory.xlsx",
        )
    flash("不支持的导出格式", "error")
    return redirect(url_for("manage"))


@app.route("/members")
def members():
    members = load_members()
    return render_template("members.html", members=members)


@app.route("/members/add", methods=["POST"])
def members_add():
    member_id = request.form.get("member_id", "").strip()
    name = request.form.get("name", "").strip()
    discount = request.form.get("discount", "").strip()
    if not member_id or not name or not discount:
        flash("请填写完整会员信息", "error")
        return redirect(url_for("members"))
    try:
        discount = int(discount)
    except ValueError:
        flash("折扣必须是整数", "error")
        return redirect(url_for("members"))
    if discount < 0 or discount > 100:
        flash("折扣必须在 0 到 100 之间", "error")
        return redirect(url_for("members"))
    member = {"member_id": member_id, "name": name, "discount": discount}
    update_member(member)
    flash("会员已保存", "success")
    return redirect(url_for("members"))


@app.route("/members/delete/<member_id>", methods=["POST"])
def members_delete(member_id):
    member = Member.query.get(member_id)
    if member:
        db.session.delete(member)
        db.session.commit()
    flash("会员已删除", "success")
    return redirect(url_for("members"))


@app.route("/api/member/<member_id>")
def api_member(member_id):
    member = find_member(member_id)
    if member:
        return jsonify({"found": True, "member": member.to_dict()})
    return jsonify({"found": False, "error": "未找到会员"}), 404


@app.route("/cashier")
def cashier():
    items, total = cart_items()
    return render_template("cashier.html", items=items, total=total)


@app.route("/customer-display")
def customer_display():
    return render_template("customer_display.html")


@app.route("/api/customer-display")
def api_customer_display():
    return jsonify(get_customer_display_payload())


@app.route("/sales")
def sales():
    report_date = parse_report_date(request.args.get("date", ""))
    sales = sales_for_date(report_date)
    rows = sale_rows(sales)
    summary = {
        "orders": len(sales),
        "items": sum(row["qty"] for row in rows),
        "total": round(sum(sale.total for sale in sales)),
        "payable": round(sum(sale.payable for sale in sales)),
    }
    return render_template(
        "sales.html",
        sales=sales,
        rows=rows,
        summary=summary,
        report_date=report_date.strftime("%Y-%m-%d"),
    )


@app.route("/sales/export/<fmt>")
def sales_export(fmt):
    report_date = parse_report_date(request.args.get("date", ""))
    sales = sales_for_date(report_date)
    rows = sale_rows(sales)
    headers = [
        "sale_id", "created_at", "payment_method", "member_id", "member_name", "discount",
        "barcode", "name", "price", "qty", "subtotal", "order_total", "order_payable",
    ]
    filename_date = report_date.strftime("%Y%m%d")

    if fmt == "csv":
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)
        data = output.getvalue().encode("utf-8-sig")
        return send_file(
            io.BytesIO(data),
            mimetype="text/csv",
            as_attachment=True,
            download_name=f"sales-{filename_date}.csv",
        )

    if fmt == "xlsx":
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Sales"
        sheet.append(headers)
        for row in rows:
            sheet.append([row.get(header, "") for header in headers])
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f"sales-{filename_date}.xlsx",
        )

    flash("不支持的导出格式", "error")
    return redirect(url_for("sales", date=report_date.strftime("%Y-%m-%d")))


@app.route("/api/product/<barcode>")
def api_product(barcode):
    product = find_product(barcode)
    if product:
        return jsonify({"found": True, "product": product.to_dict()})
    return jsonify({"found": False}), 404


@app.route("/api/products/search")
def api_products_search():
    keyword = request.args.get("q", "").strip()
    query = Product.query
    if keyword:
        pattern = f"%{keyword}%"
        query = query.filter(
            or_(
                Product.name.ilike(pattern),
                Product.barcode.ilike(pattern),
            )
        )
    products = query.order_by(Product.name).limit(24).all()
    return jsonify({
        "products": [product.to_dict() for product in products],
    })


@app.route("/api/cashier/scan", methods=["POST"])
def api_cashier_scan():
    data = request.get_json(force=True, silent=True) or {}
    barcode = str(data.get("barcode", "")).strip()
    if not barcode:
        return jsonify({"error": "barcode required"}), 400
    product = find_product(barcode)
    if not product:
        return jsonify({"error": "未找到商品"}), 404

    cart = get_cart()
    qty = cart.get(barcode, 0) + 1
    if qty > product.stock:
        return jsonify({"error": "库存不足"}), 400

    cart[barcode] = qty
    save_cart(cart)
    cart_data = cart_payload()
    save_customer_display(display_payload(cart_data["items"], cart_data["total"], "active"))
    return jsonify({
        "product": {
            "barcode": product.barcode,
            "name": product.name,
            "price": product.price,
            "stock": product.stock,
            "qty": qty,
            "image": product.image,
        },
        "cart": cart_data,
    })


@app.route("/api/cashier/cart/<barcode>", methods=["PATCH"])
def api_update_cart_item(barcode):
    data = request.get_json(force=True, silent=True) or {}
    try:
        qty = int(data.get("qty", 0))
    except (ValueError, TypeError):
        return jsonify({"error": "数量必须是整数"}), 400

    cart = get_cart()
    if barcode not in cart:
        return jsonify({"error": "购物车中没有该商品"}), 404

    if qty <= 0:
        cart.pop(barcode, None)
        save_cart(cart)
        cart_data = cart_payload()
        save_customer_display(display_payload(cart_data["items"], cart_data["total"], "active"))
        return jsonify({"success": True, "cart": cart_data})

    product = find_product(barcode)
    if not product:
        cart.pop(barcode, None)
        save_cart(cart)
        cart_data = cart_payload()
        save_customer_display(display_payload(cart_data["items"], cart_data["total"], "active"))
        return jsonify({"error": "商品不存在，已从购物车移除", "cart": cart_data}), 404

    if qty > product.stock:
        return jsonify({"error": f"库存不足，当前库存 {product.stock}"}), 400

    cart[barcode] = qty
    save_cart(cart)
    cart_data = cart_payload()
    save_customer_display(display_payload(cart_data["items"], cart_data["total"], "active"))
    return jsonify({"success": True, "cart": cart_data})


@app.route("/api/cashier/cart/<barcode>", methods=["DELETE"])
def api_delete_cart_item(barcode):
    cart = get_cart()
    cart.pop(barcode, None)
    save_cart(cart)
    cart_data = cart_payload()
    save_customer_display(display_payload(cart_data["items"], cart_data["total"], "active"))
    return jsonify({"success": True, "cart": cart_data})


@app.route("/api/cashier/checkout", methods=["POST"])
def api_checkout():
    data = request.get_json(force=True, silent=True) or {}
    member_id = str(data.get("member_id", "")).strip()
    payment_method = normalize_payment_method(data.get("payment_method", "移动支付"))
    member = find_member(member_id) if member_id else None
    cart = get_cart()
    if not cart:
        return jsonify({"error": "购物车为空"}), 400

    checkout_items = []
    total = 0.0
    for barcode, qty in cart.items():
        product = find_product(barcode)
        if not product:
            return jsonify({"error": f"商品 {barcode} 不存在，请先移出购物车"}), 400
        if qty > product.stock:
            return jsonify({"error": f"{product.name} 库存不足，当前库存 {product.stock}"}), 400
        subtotal = round(product.price * qty)
        total += subtotal
        checkout_items.append({
            "product": product,
            "barcode": product.barcode,
            "name": product.name,
            "price": product.price,
            "qty": qty,
            "subtotal": subtotal,
            "image": product.image,
        })

    total = round(total)
    payable = calculate_payable(total, member)
    sale = Sale(
        created_at=datetime.now(),
        member_id=member.member_id if member else None,
        member_name=member.name if member else None,
        discount=member.discount if member else 0,
        payment_method=payment_method,
        total=total,
        payable=payable,
    )
    db.session.add(sale)

    for item in checkout_items:
        product = item["product"]
        qty = item["qty"]
        product.stock -= qty
        db.session.add(SaleItem(
            sale=sale,
            barcode=item["barcode"],
            name=item["name"],
            price=item["price"],
            qty=qty,
            subtotal=item["subtotal"],
        ))
    db.session.commit()
    save_customer_display(display_payload(
        [{key: item[key] for key in ("barcode", "name", "price", "qty", "subtotal", "image")} for item in checkout_items],
        payable,
        "paid",
        checkout_id=sale.id,
    ))
    session["cart"] = {}
    message = f"结账完成，已记录销售单 #{sale.id}（{payment_method}）"
    if member:
        message += f"（{member.name} 享受 {member.discount}% 折扣）"
    return jsonify({"success": True, "message": message, "sale_id": sale.id})


@app.route("/api/cashier/cancel", methods=["POST"])
def api_cancel():
    session["cart"] = {}
    save_customer_display(display_payload(status="idle"))
    return jsonify({"success": True})


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    debug = os.environ.get("FLASK_DEBUG", "false").lower() in {"1", "true", "yes"}
    app.run(host="0.0.0.0", port=port, debug=debug)
